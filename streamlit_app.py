import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from openpyxl.styles import PatternFill

from logic.file_loader import (
    load_file,
    apply_column_kinds,
    infer_column_kinds,
    KIND_NUMERIC, KIND_CATEGORICAL, KIND_DATETIME, KIND_ID, KIND_EXCLUDED,
    ALL_KINDS,
)
from logic.preprocessing import (
    detect_column_types,
    analyze_column_quality,
    detect_id_columns,
    get_safe_columns,
    detect_column_groups,
    propose_balanced_selection,
    preprocess,
)
from logic.anomaly_model import (
    benchmark_models,
    compute_ensemble,
    get_top_deviating_features,
    explain_record,
    detect_temporal_anomalies,
    run_semi_supervised,
    MODEL_DESCRIPTIONS,
    MODELS,
)
from logic.data_quality import run_quality_checks, severity_label
from logic.presets import PRESETS
from logic.feedback_store import (
    dataset_signature,
    save_feedback,
    load_feedback_for,
    derive_model_weights,
)

NEON_YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def to_styled_excel(dataframe: pd.DataFrame, anomaly_threshold: float) -> bytes:
    """Write DataFrame to Excel with neon yellow highlight on anomaly rows."""
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="Sonuçlar")
        ws = writer.sheets["Sonuçlar"]

        score_col_idx = None
        for idx, col in enumerate(dataframe.columns, start=1):
            if col == "anomaly_score":
                score_col_idx = idx
                break

        if score_col_idx is not None:
            for row_idx in range(2, len(dataframe) + 2):
                cell_value = ws.cell(row=row_idx, column=score_col_idx).value
                if cell_value is not None and cell_value >= anomaly_threshold:
                    for col_idx in range(1, len(dataframe.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = NEON_YELLOW

    return output.getvalue()


def _invalidate_pipeline_state():
    """Clear cached pipeline outputs so they re-run with fresh inputs."""
    for k in (
        "bench_results", "best_key", "ensemble", "result_df", "scores",
        "processed", "feature_names", "active_view",
    ):
        st.session_state.pop(k, None)


# ── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(page_title="Anomaly Detection", page_icon="🔍", layout="wide")
st.title("Anomaly Detection Agent")
st.markdown("Veri setinizi yükleyin, sütunları seçin ve olağandışı kayıtları tespit edin.")

# ── Sidebar — Settings + Tips ─────────────────────────────────────────────────
st.sidebar.header("Ayarlar")

# Preset is read first because it influences contamination range and weights.
preset_name = st.sidebar.selectbox("Veri seti türünüz", list(PRESETS.keys()))
preset = PRESETS[preset_name]

cmin, cmax = preset.get("contamination_range", (0.01, 0.30))
contamination = st.sidebar.slider(
    "Beklenen Anomali Oranı",
    min_value=float(cmin),
    max_value=float(cmax),
    value=float(preset.get("contamination_default", 0.05)),
    step=0.01,
    help="Veri setinde yaklaşık ne kadar olağandışı kayıt bekliyorsunuz?",
)

top_n = st.sidebar.slider(
    "Gösterilecek Şüpheli Kayıt Sayısı",
    min_value=5,
    max_value=100,
    value=20,
    step=5,
)

st.sidebar.divider()
st.sidebar.header("İpuçları")
tip_text = f"**{preset['description']}**\n\n"
for tip in preset["tips"]:
    tip_text += f"- {tip}\n"
st.sidebar.info(tip_text)

# ── 1. File upload ───────────────────────────────────────────────────────────
st.header("1. Veri Yükleme")
uploaded_file = st.file_uploader(
    "CSV, Excel, XML veya ZIP dosyası yükleyin",
    type=["csv", "xls", "xlsx", "xml", "zip"],
)

if uploaded_file is None:
    st.info("Başlamak için bir dosya yükleyin.")
    st.stop()

# Cache the raw read on the (name, size) signature to avoid re-parsing on every rerun.
file_sig = (uploaded_file.name, getattr(uploaded_file, "size", None))
if st.session_state.get("file_sig") != file_sig:
    df_raw = load_file(uploaded_file)
    if df_raw is None:
        st.stop()
    st.session_state["df_raw"] = df_raw
    st.session_state["file_sig"] = file_sig
    st.session_state.pop("kinds", None)
    st.session_state.pop("df", None)
    st.session_state.pop("kinds_report", None)
    _invalidate_pipeline_state()

df_raw: pd.DataFrame = st.session_state["df_raw"]

# ── 2. Data preview ──────────────────────────────────────────────────────────
st.header("2. Veri Önizleme")

col1, col2, col3 = st.columns(3)
col1.metric("Satır Sayısı", df_raw.shape[0])
col2.metric("Sütun Sayısı", df_raw.shape[1])
col3.metric("Eksik Değer", int(df_raw.isna().sum().sum()))

with st.expander("İlk 100 satırı göster", expanded=False):
    st.dataframe(df_raw.head(100), use_container_width=True)

# ── 3. Kolon Tipi Kontrolü (NEW) ────────────────────────────────────────────
st.header("3. Kolon Tipi Kontrolü")
st.markdown(
    "Sistem her kolonu otomatik tipler. Yanlış algılanan kolonu açılır listeden değiştirdikten sonra "
    "**Tipleri Uygula**'ya basın. "
)

# Auto-infer kinds (with preset numeric_hints applied by extending name match)
def _augmented_infer(df: pd.DataFrame, preset_hints: tuple[str, ...]) -> dict:
    base = infer_column_kinds(df)
    if preset_hints:
        for col in df.columns:
            cl = str(col).lower()
            if base.get(col) == KIND_CATEGORICAL and any(h in cl for h in preset_hints):
                base[col] = KIND_NUMERIC
    return base


if "kinds" not in st.session_state:
    st.session_state["kinds"] = _augmented_infer(df_raw, preset.get("numeric_hints", ()))

kinds_df = pd.DataFrame({
    "Sütun": list(df_raw.columns),
    "Otomatik Tip": [st.session_state["kinds"].get(c, KIND_CATEGORICAL) for c in df_raw.columns],
    "Örnek Değer": [
        str(df_raw[c].dropna().iloc[0]) if df_raw[c].notna().any() else ""
        for c in df_raw.columns
    ],
    "Eksik %": [
        round(df_raw[c].isna().sum() / max(len(df_raw), 1) * 100, 1)
        for c in df_raw.columns
    ],
})

edited = st.data_editor(
    kinds_df,
    use_container_width=True,
    hide_index=True,
    num_rows="fixed",
    key="kinds_editor",
    column_config={
        "Sütun": st.column_config.TextColumn(disabled=True),
        "Otomatik Tip": st.column_config.SelectboxColumn(
            "Tip",
            options=list(ALL_KINDS),
            required=True,
            help="Tipi yanlış algılandıysa düzeltin",
        ),
        "Örnek Değer": st.column_config.TextColumn(disabled=True),
        "Eksik %": st.column_config.NumberColumn(disabled=True, format="%.1f"),
    },
)

apply_btn_col, reset_btn_col, _ = st.columns([1, 1, 4])
apply_clicked = apply_btn_col.button("Tipleri Uygula", type="primary", use_container_width=True)
reset_clicked = reset_btn_col.button("Otomatik Tipe Dön", use_container_width=True)

if reset_clicked:
    st.session_state["kinds"] = _augmented_infer(df_raw, preset.get("numeric_hints", ()))
    st.session_state.pop("df", None)
    st.session_state.pop("kinds_report", None)
    _invalidate_pipeline_state()
    st.rerun() if hasattr(st, "rerun") else st.experimental_rerun()

if apply_clicked or "df" not in st.session_state:
    new_kinds = {row["Sütun"]: row["Otomatik Tip"] for _, row in edited.iterrows()}
    st.session_state["kinds"] = new_kinds
    df_typed, kinds_report = apply_column_kinds(df_raw, new_kinds)
    st.session_state["df"] = df_typed
    st.session_state["kinds_report"] = kinds_report
    _invalidate_pipeline_state()

df: pd.DataFrame = st.session_state["df"]
kinds: dict = st.session_state["kinds"]
kinds_report: dict = st.session_state.get("kinds_report", {})

# Conversion report
if kinds_report:
    converted_rows = []
    for col, info in kinds_report.items():
        if info.get("already_numeric") or info.get("already_datetime"):
            continue
        target = "Sayısal" if "non_null_original" in info and "failed" in info and not info.get("already_datetime") else "Tarih"
        if kinds.get(col) == KIND_DATETIME:
            target = "Tarih"
        elif kinds.get(col) == KIND_NUMERIC:
            target = "Sayısal"
        converted_rows.append({
            "Sütun": col,
            "Hedef": target,
            "Çevrilen": info.get("converted", 0),
            "Çevrilemeyen": info.get("failed", 0),
            "Örnek Çevrilemeyen": ", ".join(info.get("failed_samples", [])[:3]),
        })
    if converted_rows:
        with st.expander("Tip Dönüşüm Raporu", expanded=any(r["Çevrilemeyen"] > 0 for r in converted_rows)):
            st.dataframe(pd.DataFrame(converted_rows), use_container_width=True, hide_index=True)
            n_failed = sum(r["Çevrilemeyen"] for r in converted_rows)
            if n_failed > 0:
                st.warning(
                    f"Toplam **{n_failed}** değer hedef tipe çevrilemedi. "
                    f"Bu hücreler boş (`NaN`) olarak işaretlendi, modele eksik veri olarak girer."
                )
            else:
                st.success("Tüm değerler hedef tipe başarıyla çevrildi.")

# ── 4. Kolon Kalite Kontrolü ────────────────────────────────────────────────
st.header("4. Kolon Kalite Kontrolü")
quality_df = analyze_column_quality(df)
flagged = quality_df[quality_df["flags"] != "-"]

if not flagged.empty:
    st.info(
        f"{len(flagged)} sütunda öneri tespit edildi. "
        f"Bu sütunları yine de kullanabilirsiniz, ancak sonuçları etkileyebilir."
    )
    with st.expander("Kalite önerileri (detay)", expanded=False):
        flag_display = flagged[["column", "dtype", "unique", "unique_ratio", "missing_pct", "flags"]].copy()
        flag_display.columns = ["Sütun", "Tip", "Benzersiz", "Benzersiz Oranı", "Eksik %", "Öneri"]
        st.dataframe(flag_display, use_container_width=True, hide_index=True)
else:
    st.success("Tüm sütunlar kalite kontrolünden geçti.")

# ── 5. Tanımlayıcı (ID) Seçimi ──────────────────────────────────────────────
st.header("5. Tanımlayıcı (ID) Seçimi")

id_candidates = detect_id_columns(df, quality_df)
# Also include columns user marked as KIND_ID
id_candidates = sorted(set(id_candidates) | {c for c, k in kinds.items() if k == KIND_ID})

if id_candidates:
    st.markdown(
        f"Veri setinizde **{len(id_candidates)}** adet tanımlayıcı olabilecek sütun var: "
        f"**{', '.join(id_candidates)}**"
    )
    selected_ids = st.multiselect(
        "Tanımlayıcı olarak kullanılacak sütunlar",
        options=id_candidates,
        default=id_candidates,
    )
else:
    st.success("Tanımlayıcı tespit edilmedi.")
    selected_ids = []

# ── 6. Kolon Kapsamı Kontrolü (NEW) ─────────────────────────────────────────
st.header("6. Kolon Kapsamı Kontrolü")

# Excluded explicit by user (KIND_EXCLUDED) → never enter analysis
user_excluded = [c for c, k in kinds.items() if k == KIND_EXCLUDED]

# Build balanced default selection
proposal_cols, coverage_report = propose_balanced_selection(
    df=df.drop(columns=user_excluded, errors="ignore"),
    quality_df=analyze_column_quality(df.drop(columns=user_excluded, errors="ignore")),
    id_cols=selected_ids,
    max_columns=30,
)

usable_pool = [c for c in df.columns if c not in user_excluded]

st.markdown(
    "Sistem, sayısal alanlar + düşük/orta kardinaliteli kategorik alanları dengeli olarak seçer. "
    "Aşağıda hangi kolonların **dahil**, hangilerinin **dışarıda** olduğunu ve nedenini görebilirsiniz."
)

cov_left, cov_right = st.columns(2)
with cov_left:
    st.markdown("**Default kapsam (dahil)**")
    inc_df = pd.DataFrame([r for r in coverage_report if r["included"]])
    if not inc_df.empty:
        inc_df = inc_df.rename(columns={"column": "Sütun", "reason": "Neden dahil"})
        st.dataframe(inc_df[["Sütun", "Neden dahil"]], use_container_width=True, hide_index=True)
    else:
        st.info("Default kapsamda kolon yok.")

with cov_right:
    st.markdown("**Dışarıda kalan kolonlar**")
    exc_df = pd.DataFrame([r for r in coverage_report if not r["included"]])
    if not exc_df.empty:
        exc_df = exc_df.rename(columns={"column": "Sütun", "reason": "Neden hariç"})
        st.dataframe(exc_df[["Sütun", "Neden hariç"]], use_container_width=True, hide_index=True)
    else:
        st.success("Tüm kolonlar dahil edildi.")

# Group view
groups = detect_column_groups(usable_pool)
if len(groups) > 1:
    with st.expander("Kolonları gruplara göre incele", expanded=False):
        for prefix, cols in groups.items():
            st.markdown(f"**{prefix}** ({len(cols)} kolon): {', '.join(cols)}")

selected_columns = st.multiselect(
    "Analiz edilecek sütunları seçin (default: dengeli kapsam)",
    options=usable_pool,
    default=[c for c in proposal_cols if c not in selected_ids],
    help="ID olarak seçilen kolonlar burada listelenmez.",
)

if not selected_columns:
    st.warning("En az bir sütun seçmelisiniz.")
    st.stop()

# Optional: combo column picker for joint-frequency feature
combo_eligible = [
    c for c in selected_columns
    if not pd.api.types.is_numeric_dtype(df[c]) and df[c].nunique() <= 200
]
combo_columns = st.multiselect(
    "Beklenmedik kombinasyon kontrolü için kolonlar (opsiyonel, ≥2 seçin)",
    options=combo_eligible,
    default=[],
    help=(
        "Örn. para birimi + kanal + işlem tipi: bu üçlü veri setinde nadir ise modele "
        "ek bir 'rare combo' sinyali olarak girer."
    ),
)

# ── 7. Veri Kalitesi Kontrolleri (NEW) ──────────────────────────────────────
st.header("7. Veri Kalitesi Kontrolleri")
st.markdown(
    "Bu kontroller anomali tespitinden önce bağımsız bir şekilde çalışır. İstatistiksel olarak "
    "format/içerik bozukluklarını yakalar."
)

inject_dq = st.toggle(
    "Veri kalitesi kontrollerini modele ek özellik olarak da ekle",
    value=True,
    help="Açık tutulduğunda, kural ihlali işaretleri preprocess sırasında binary feature olarak modele girer.",
)

dq_summary, dq_flags = run_quality_checks(
    df,
    kinds=kinds,
    allowed_currencies=preset.get("allowed_currencies", []),
    required_columns=[],
)

# Filter out checks the preset disabled (they're computed but hidden + not injected)
preset_rule_toggles = preset.get("rule_checks", {})
if not dq_summary.empty:
    dq_summary_visible = dq_summary[
        dq_summary["rule"].apply(lambda r: preset_rule_toggles.get(r, True))
    ].copy()
else:
    dq_summary_visible = dq_summary

if dq_summary_visible.empty:
    st.success("Veri kalitesi kontrollerinden bayrak çıkmadı.")
    dq_flags_for_features = pd.DataFrame(index=df.index)
else:
    dq_summary_visible["Önem"] = dq_summary_visible["severity"].apply(severity_label)
    dq_summary_visible["Örnekler"] = dq_summary_visible["samples"].apply(
        lambda xs: ", ".join(map(str, xs)) if isinstance(xs, list) else str(xs)
    )
    show = dq_summary_visible.rename(columns={
        "rule": "Kural", "column": "Sütun", "count": "Adet",
        "share": "Oran %", "description": "Açıklama",
    })[["Önem", "Kural", "Sütun", "Adet", "Oran %", "Açıklama", "Örnekler"]]
    st.dataframe(show, use_container_width=True, hide_index=True)

    # Filter row flags to just enabled rules
    enabled_flag_cols = [
        c for c in dq_flags.columns
        if preset_rule_toggles.get(c.split("__")[0], True)
    ]
    dq_flags_for_features = dq_flags[enabled_flag_cols] if enabled_flag_cols else pd.DataFrame(index=df.index)

# ── 8. Anomali Tespiti ──────────────────────────────────────────────────────
st.header("8. Anomali Tespiti")
st.markdown(
    "Sistem üç farklı yöntemi otomatik olarak çalıştırır, karşılaştırır ve toplu bir "
    "sonuç üretir. Tek bir 'en iyi model' seçimi yerine her modelin sonuçlarını ayrı sekmelerde "
    "incelemeniz mümkündür."
)

if st.button("Analizi Başlat", type="primary", use_container_width=True):
    with st.spinner("Veri hazırlanıyor ve modeller çalıştırılıyor..."):
        encoding = preset.get("encoding", {})
        processed, scaler, feature_names = preprocess(
            df,
            selected_columns,
            max_onehot_categories=encoding.get("onehot_max_unique", 15),
            rare_category_columns=[
                c for c in selected_columns
                if not pd.api.types.is_numeric_dtype(df[c]) and df[c].nunique() <= 200
            ],
            combo_columns=combo_columns if len(combo_columns) >= 2 else None,
            add_rule_features=False,  # rule features come via dq_flags injection
            extra_features=dq_flags_for_features if inject_dq else None,
        )

        # Adjust model weights from preset + persistent feedback
        base_weights = preset.get("model_weights", {})
        adjusted_weights = derive_model_weights(preset_name, base_weights)

        bench_results, best_key = benchmark_models(
            processed,
            contamination=contamination,
            model_weights=adjusted_weights,
        )
        ensemble = compute_ensemble(
            bench_results,
            contamination=contamination,
            model_weights=adjusted_weights,
        )

    st.session_state["processed"] = processed
    st.session_state["feature_names"] = feature_names
    st.session_state["bench_results"] = bench_results
    st.session_state["best_key"] = best_key
    st.session_state["ensemble"] = ensemble
    st.session_state["selected_columns"] = selected_columns
    st.session_state["selected_ids"] = selected_ids
    st.session_state["combo_columns"] = combo_columns
    st.session_state["dq_flags"] = dq_flags_for_features
    st.session_state["adjusted_weights"] = adjusted_weights

    if "feedback" not in st.session_state:
        # Try to load any prior persisted feedback for this dataset+columns combo
        sig = dataset_signature(uploaded_file.name, len(df), selected_columns)
        prior = load_feedback_for(sig)
        st.session_state["feedback"] = {e["row_idx"]: e["label"] for e in prior}
        st.session_state["dataset_signature"] = sig

if "bench_results" not in st.session_state:
    st.stop()

bench_results = st.session_state["bench_results"]
if not any(r.get("success") for r in bench_results):
    st.error("Üç modelin de eğitimi başarısız oldu. Kolon seçimi ve veri tiplerini kontrol edin.")
    for r in bench_results:
        if not r.get("success"):
            st.caption(f"• {r['model_key']}: {r.get('error', 'bilinmeyen hata')}")
    st.stop()

best_key = st.session_state["best_key"]
ensemble = st.session_state["ensemble"]
processed = st.session_state["processed"]
sel_columns = st.session_state["selected_columns"]
sel_ids = st.session_state.get("selected_ids", [])
combo_cols = st.session_state.get("combo_columns", [])
dq_flags_state = st.session_state.get("dq_flags", pd.DataFrame(index=df.index))

# ── 9. Model Karşılaştırması ────────────────────────────────────────────────
st.header("9. Model Karşılaştırması")

best_display = [k for k, v in MODELS.items() if v == best_key][0]
st.success(f"Ağırlıklı uyum skoruna göre öne çıkan model: **{best_display}**")

bench_rows = []
for r in bench_results:
    display_name = [k for k, v in MODELS.items() if v == r["model_key"]][0]
    bench_rows.append({
        "Model": display_name,
        "Açıklama": MODEL_DESCRIPTIONS[r["model_key"]],
        "Bulunan Anomali": r["n_anomalies"] if r["success"] else "-",
        "Uyum Skoru (silhouette)": r["silhouette"] if r["success"] else "-",
        "Preset Ağırlığı": st.session_state.get("adjusted_weights", {}).get(r["model_key"], 1.0),
    })
st.dataframe(pd.DataFrame(bench_rows), use_container_width=True, hide_index=True)

# Build a per-model result_df dict so the tabs below can switch view smoothly.
model_views: dict[str, pd.DataFrame] = {}
for r in bench_results:
    if not r["success"]:
        continue
    mk = r["model_key"]
    scores_arr = pd.Series(r["scores"], index=processed.index, name="anomaly_score")
    rdf = df.copy()
    rdf["anomaly_score"] = scores_arr
    rdf = rdf.sort_values("anomaly_score", ascending=False)
    rdf["rank"] = range(1, len(rdf) + 1)
    model_views[mk] = rdf

if ensemble.get("ensemble_score") is not None:
    ens_scores = pd.Series(ensemble["ensemble_score"], index=processed.index, name="anomaly_score")
    edf = df.copy()
    edf["anomaly_score"] = ens_scores
    if ensemble.get("any_top_mask") is not None:
        edf["any_top_in_models"] = ensemble["any_top_mask"]
    edf = edf.sort_values("anomaly_score", ascending=False)
    edf["rank"] = range(1, len(edf) + 1)
    model_views["ensemble"] = edf

# ── 10. Sonuçlar (tabs per model + ensemble) ────────────────────────────────
st.header("10. Sonuçlar")

tab_labels = []
tab_keys = []
if "ensemble" in model_views:
    tab_labels.append("Önerilen")
    tab_keys.append("ensemble")
for r in bench_results:
    if r["success"]:
        tab_labels.append([k for k, v in MODELS.items() if v == r["model_key"]][0])
        tab_keys.append(r["model_key"])

tabs = st.tabs(tab_labels)


def render_view(view_df: pd.DataFrame, view_key: str, top_n_local: int) -> None:
    if view_key == "ensemble":
        threshold_local = view_df["anomaly_score"].quantile(1 - contamination)
    else:
        threshold_local = view_df["anomaly_score"].quantile(1 - contamination)
    n_anom_local = int((view_df["anomaly_score"] >= threshold_local).sum())

    m1, m2, m3 = st.columns(3)
    m1.metric("Toplam Kayıt", len(view_df))
    m2.metric("Şüpheli Kayıt", n_anom_local)
    m3.metric("Şüpheli Oranı", f"%{n_anom_local / max(len(view_df), 1) * 100:.1f}")

    if view_key == "ensemble" and "any_top_in_models" in view_df.columns:
        any_count = int(view_df["any_top_in_models"].sum())
        st.caption(
            f"En az **bir** modelin top-%{contamination * 100:.0f}'ine giren kayıt sayısı: **{any_count}**"
        )

    fig_local = px.histogram(
        view_df,
        x="anomaly_score",
        nbins=50,
        labels={"anomaly_score": "Şüphelilik Puanı", "count": "Kayıt Sayısı"},
    )
    fig_local.add_vline(x=threshold_local, line_dash="dash", line_color="red", annotation_text="Eşik")
    fig_local.update_layout(title=None)
    st.plotly_chart(fig_local, use_container_width=True)

    top = view_df.head(top_n_local)
    if sel_ids:
        id_cols_present = [c for c in sel_ids if c in top.columns]
        other_cols = [c for c in top.columns if c not in id_cols_present]
        ordered = id_cols_present + ["anomaly_score", "rank"] + [
            c for c in other_cols if c not in ["anomaly_score", "rank"]
        ]
        st.dataframe(
            top[ordered].style.background_gradient(subset=["anomaly_score"], cmap="Reds"),
            use_container_width=True,
            hide_index=True,
        )
    else:
        st.dataframe(
            top.style.background_gradient(subset=["anomaly_score"], cmap="Reds"),
            use_container_width=True,
            hide_index=True,
        )


for tab, key in zip(tabs, tab_keys):
    with tab:
        render_view(model_views[key], key, top_n)

# Pick the active result set for downstream sections.
active_key = "ensemble" if "ensemble" in model_views else best_key
result_df = model_views[active_key]
scores = result_df["anomaly_score"]
threshold = scores.quantile(1 - contamination)
n_anomalies = int((scores >= threshold).sum())
st.session_state["result_df"] = result_df
st.session_state["scores"] = scores

# ── 11. Kayıt İnceleme ──────────────────────────────────────────────────────
st.header("11. Kayıt İnceleme")
st.markdown("Aşağıdaki listeden bir kaydı seçerek **neden şüpheli bulunduğunu** inceleyin.")

top_anomalies = result_df.head(top_n)


def _format_rank(x):
    r = top_anomalies.loc[top_anomalies["rank"] == x].iloc[0]
    id_parts = [f"{c}: {r[c]}" for c in sel_ids if c in r.index] if sel_ids else []
    id_label = f" ({', '.join(id_parts)})" if id_parts else ""
    return f"Sıra {x}{id_label} — Şüphelilik: {r['anomaly_score']:.3f}"


selected_rank = st.selectbox(
    "İncelemek istediğiniz kaydı seçin",
    options=top_anomalies["rank"].tolist(),
    format_func=_format_rank,
)

row_mask = top_anomalies["rank"] == selected_rank
original_idx = top_anomalies.loc[row_mask].index[0]
row = top_anomalies.loc[row_mask].iloc[0]

detail_col, reason_col = st.columns(2)

with detail_col:
    st.subheader("Kayıt Bilgileri")
    if sel_ids:
        id_values = {c: row[c] for c in sel_ids if c in row.index}
        if id_values:
            id_text = " &nbsp;|&nbsp; ".join([f"**{k}:** {v}" for k, v in id_values.items()])
            st.markdown(id_text)
            st.divider()
    drop_cols = ["anomaly_score", "rank"]
    if "any_top_in_models" in row.index:
        drop_cols.append("any_top_in_models")
    row_dict = row.drop(drop_cols, errors="ignore").to_dict()
    display_items = [{"Alan": k, "Değer": v} for k, v in row_dict.items()]
    st.dataframe(pd.DataFrame(display_items), use_container_width=True, hide_index=True)

with reason_col:
    st.subheader("Neden Şüpheli?")
    flags_row = (
        dq_flags_state.loc[original_idx]
        if (not dq_flags_state.empty and original_idx in dq_flags_state.index)
        else None
    )
    reasons = explain_record(
        df.loc[original_idx],
        df,
        sel_columns,
        rare_threshold=preset.get("encoding", {}).get("rare_threshold", 0.005),
        combo_columns=combo_cols if len(combo_cols) >= 2 else None,
        dq_flags_row=flags_row,
        top_k=8,
    )
    if not reasons:
        st.info("Bu kayıt için anlamlı bir sapma sinyali bulunamadı, model komşuluk yapısı üzerinden işaretlemiş olabilir.")
    else:
        groups_by_kind = {"numeric": [], "kategorik": [], "kombinasyon": [], "kural": []}
        for r in reasons:
            groups_by_kind.setdefault(r["kind"], []).append(r)
        kind_titles = {
            "numeric": "Sayısal sapmalar",
            "kategorik": "Nadir kategorik değerler",
            "kombinasyon": "Olağan dışı kombinasyon",
            "kural": "Kural ihlalleri",
        }
        for k, items in groups_by_kind.items():
            if not items:
                continue
            st.markdown(f"**{kind_titles[k]}**")
            for it in items:
                st.markdown(f"- {it['message']}")

# ── 12. Zaman Serisi Analizi (existing) ─────────────────────────────────────
col_types = detect_column_types(df)
datetime_cols = col_types["datetime"]
numeric_cols_for_ts = [c for c in sel_columns if c in col_types["numeric"]]

if datetime_cols and numeric_cols_for_ts:
    st.header("12. Zaman Serisi Analizi")
    st.markdown(
        "Verilerinizde tarih sütunu tespit edildi. Zaman içindeki olağandışı değişimleri "
        "aşağıda inceleyebilirsiniz."
    )

    ts_col1, ts_col2 = st.columns(2)
    with ts_col1:
        ts_date_col = st.selectbox("Tarih sütunu", datetime_cols)
    with ts_col2:
        ts_value_col = st.selectbox("İncelenecek değer sütunu", numeric_cols_for_ts)

    ts_window = st.slider("Hareketli ortalama penceresi (gün)", 3, 60, 14)

    if st.button("Zaman Serisi Analizi Çalıştır", use_container_width=True):
        with st.spinner("Zaman serisi analiz ediliyor..."):
            ts_result = detect_temporal_anomalies(df, ts_date_col, ts_value_col, window=ts_window)

        n_ts_anomalies = int(ts_result["is_anomaly"].sum())
        st.metric("Zaman Serisinde Şüpheli Nokta", n_ts_anomalies)

        fig_ts = go.Figure()
        normal_ts = ts_result[~ts_result["is_anomaly"]]
        fig_ts.add_trace(go.Scatter(
            x=normal_ts[ts_date_col], y=normal_ts[ts_value_col],
            mode="markers", name="Normal", marker=dict(color="#636EFA", size=4),
        ))
        anom_ts = ts_result[ts_result["is_anomaly"]]
        fig_ts.add_trace(go.Scatter(
            x=anom_ts[ts_date_col], y=anom_ts[ts_value_col],
            mode="markers", name="Şüpheli", marker=dict(color="#EF553B", size=10, symbol="x"),
        ))
        fig_ts.add_trace(go.Scatter(
            x=ts_result[ts_date_col], y=ts_result["rolling_mean"],
            mode="lines", name="Ortalama Trend", line=dict(color="gray", dash="dash"),
        ))
        fig_ts.add_trace(go.Scatter(
            x=ts_result[ts_date_col], y=ts_result["upper"],
            mode="lines", name="Üst Sınır", line=dict(color="rgba(200,200,200,0.5)"),
        ))
        fig_ts.add_trace(go.Scatter(
            x=ts_result[ts_date_col], y=ts_result["lower"],
            mode="lines", name="Alt Sınır", line=dict(color="rgba(200,200,200,0.5)"),
            fill="tonexty", fillcolor="rgba(200,200,200,0.15)",
        ))
        fig_ts.update_layout(
            xaxis_title="Tarih",
            yaxis_title=ts_value_col,
            height=450,
            legend=dict(orientation="h", yanchor="bottom", y=1.02),
        )
        st.plotly_chart(fig_ts, use_container_width=True)

    section_offset = 13
else:
    section_offset = 12

# ── Feedback loop ────────────────────────────────────────────────────────────
st.header(f"{section_offset}. Geri Bildirim")
st.markdown(
    "Tespit edilen kayıtları **gerçek anomali** veya **yanlış alarm** olarak işaretleyin. "
    "Geri bildirimler sonraki çalışmalarda preset model ağırlıklarının ayarlanmasında kullanılır."
)

if "feedback" not in st.session_state:
    st.session_state["feedback"] = {}


def _mark_feedback(record_idx, label_value: int) -> None:
    st.session_state["feedback"][record_idx] = label_value


feedback_ranks = top_anomalies["rank"].tolist()
for rank_val in feedback_ranks[:top_n]:
    idx = top_anomalies.loc[top_anomalies["rank"] == rank_val].index[0]
    score_val = top_anomalies.loc[top_anomalies["rank"] == rank_val, "anomaly_score"].values[0]

    col_label, col_btn1, col_btn2, col_status = st.columns([3, 2, 2, 2])
    col_label.markdown(f"**Sıra {rank_val}** — Puan: {score_val:.3f}")

    current_fb = st.session_state["feedback"].get(idx, None)

    col_btn1.button(
        "Gerçek Anomali",
        key=f"anom_{idx}",
        on_click=_mark_feedback,
        args=(idx, 1),
    )
    col_btn2.button(
        "Yanlış Alarm",
        key=f"normal_{idx}",
        on_click=_mark_feedback,
        args=(idx, 0),
    )

    if current_fb == 1:
        col_status.success("Anomali")
    elif current_fb == 0:
        col_status.info("Normal")
    else:
        col_status.markdown("_Bekliyor_")

fb = st.session_state["feedback"]
n_feedback = len(fb)
n_marked_anomaly = sum(1 for v in fb.values() if v == 1)
n_marked_normal = sum(1 for v in fb.values() if v == 0)

st.divider()
fb1, fb2, fb3 = st.columns(3)
fb1.metric("Toplam Geri Bildirim", n_feedback)
fb2.metric("Gerçek Anomali", n_marked_anomaly)
fb3.metric("Yanlış Alarm", n_marked_normal)

persist_col, _ = st.columns([1, 3])
if persist_col.button("Geri Bildirimleri Kaydet", help="Bu cihazdaki kalıcı feedback dosyasına yazar"):
    sig = st.session_state.get("dataset_signature") or dataset_signature(
        uploaded_file.name, len(df), sel_columns
    )
    entries = []
    per_model_pct = ensemble.get("per_model_pct", {}) if isinstance(ensemble, dict) else {}
    ens_arr = ensemble.get("ensemble_score") if isinstance(ensemble, dict) else None
    for idx_key, label_val in fb.items():
        scores_for_row = {}
        if ens_arr is not None and idx_key in processed.index:
            pos = processed.index.get_loc(idx_key)
            scores_for_row["ensemble"] = float(ens_arr[pos])
            for mk, pct_arr in per_model_pct.items():
                scores_for_row[mk] = float(pct_arr[pos])
        entries.append({"row_idx": int(idx_key), "label": int(label_val), "scores": scores_for_row, "note": ""})
    save_feedback(
        signature=sig,
        name=uploaded_file.name,
        n_rows=len(df),
        columns_selected=sel_columns,
        preset=preset_name,
        entries=entries,
    )
    st.success(f"{len(entries)} geri bildirim kaydedildi.")

# ── Semi-supervised re-run ───────────────────────────────────────────────────
st.header(f"{section_offset + 1}. Geri Bildirimle Yeniden Analiz")
st.markdown(
    "Geri bildirimleriniz kullanılarak model güncellenir. "
    "Normal olarak işaretlediğiniz kayıtlar modele öğretilir, "
    "anomali dedikleriniz eğitimden çıkarılır."
)

min_feedback = 3
if n_feedback < min_feedback:
    st.info(f"Yeniden analiz için en az {min_feedback} geri bildirim gerekli. Şu an {n_feedback} adet var.")
else:
    if st.button("Geri Bildirimle Yeniden Analiz Et", type="primary", use_container_width=True):
        with st.spinner("Model geri bildirimlerinizle güncelleniyor..."):
            labels = pd.Series(-1, index=processed.index)
            for idx_key, label_val in fb.items():
                if idx_key in labels.index:
                    labels.loc[idx_key] = label_val

            new_scores = run_semi_supervised(processed, labels, contamination=contamination)

        result_df_v2 = df.copy()
        result_df_v2["anomaly_score"] = new_scores
        result_df_v2 = result_df_v2.sort_values("anomaly_score", ascending=False)
        result_df_v2["rank"] = range(1, len(result_df_v2) + 1)

        st.success("Model güncellendi!")

        new_threshold = new_scores.quantile(1 - contamination)
        new_n_anomalies = int((new_scores >= new_threshold).sum())

        r1, r2 = st.columns(2)
        r1.metric("Güncel Şüpheli Sayısı", new_n_anomalies, delta=new_n_anomalies - n_anomalies)
        r2.metric("Değişim", f"{abs(new_n_anomalies - n_anomalies)} kayıt")

        st.subheader(f"Güncellenmiş En Şüpheli {top_n} Kayıt")
        st.dataframe(
            result_df_v2.head(top_n).style.background_gradient(subset=["anomaly_score"], cmap="Reds"),
            use_container_width=True,
            hide_index=True,
        )

        xlsx_v2 = to_styled_excel(result_df_v2, new_threshold)
        st.download_button(
            label="Güncellenmiş Sonuçları İndir (Excel)",
            data=xlsx_v2,
            file_name="anomaly_results_v2.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

# ── Download ─────────────────────────────────────────────────────────────────
st.header(f"{section_offset + 2}. Sonuçları İndir")
st.markdown("Şüpheli bulunan kayıtlar Excel dosyasında **sarı** ile işaretlenmiştir.")

xlsx_all = to_styled_excel(result_df, threshold)
st.download_button(
    label="Tüm Sonuçları İndir (Excel)",
    data=xlsx_all,
    file_name="anomaly_results.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)

xlsx_top = to_styled_excel(top_anomalies, threshold)
st.download_button(
    label=f"En Şüpheli {top_n} Kaydı İndir (Excel)",
    data=xlsx_top,
    file_name="top_anomalies.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    use_container_width=True,
)
